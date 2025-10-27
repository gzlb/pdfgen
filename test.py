import pandas as pd
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

def read_data_with_full_style(xlsm_path):
    """Read values, styles, merged cells, and column widths from the 'DATA' sheet."""
    wb = load_workbook(xlsm_path, data_only=True)
    if "DATA" not in wb.sheetnames:
        raise ValueError(f"No DATA sheet found in {xlsm_path}")

    ws = wb["DATA"]

    # --- Extract values and cell styles ---
    values = []
    styles = []
    for row in ws.iter_rows():
        row_values = []
        row_styles = []
        for cell in row:
            row_values.append(cell.value)
            # Background color
            bg_color = None
            if cell.fill and cell.fill.start_color and cell.fill.start_color.type == 'rgb':
                bg_color = f"#{cell.fill.start_color.rgb[-6:]}"  # keep only hex color

            row_styles.append({
                "bg_color": bg_color,
                "bold": cell.font.bold,
                "italic": cell.font.italic,
                "align": (cell.alignment.horizontal or "left").upper(),
            })
        values.append(row_values)
        styles.append(row_styles)

    # --- Extract merged cell ranges ---
    merged_ranges = []
    for merged in ws.merged_cells.ranges:
        merged_ranges.append(str(merged))  # e.g. "A1:C1"

    # --- Extract column widths ---
    col_widths = []
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.width:
            # Scale Excel width (rough approximation to ReportLab units)
            col_widths.append(col_dim.width * 5)
        else:
            col_widths.append(60)

    # Fill widths if missing
    n_cols = len(values[0]) if values else 0
    if len(col_widths) < n_cols:
        col_widths += [60] * (n_cols - len(col_widths))

    wb.close()
    return values, styles, merged_ranges, col_widths


def combine_multiple_xlsm(files):
    """Combine data and style from multiple XLSM DATA sheets."""
    combined_values = []
    combined_styles = []
    combined_merges = []
    col_widths = []
    headers_added = False

    for f in files:
        values, styles, merges, widths = read_data_with_full_style(f)
        if not headers_added:
            combined_values.extend(values)
            combined_styles.extend(styles)
            combined_merges.extend(merges)
            col_widths = widths
            headers_added = True
        else:
            # skip header row for next files
            combined_values.extend(values[1:])
            combined_styles.extend(styles[1:])
            combined_merges.extend(merges)

    return combined_values, combined_styles, combined_merges, col_widths


def excel_range_to_indices(range_str):
    """Convert Excel range like 'A1:C1' to numeric coordinates."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    return (min_col - 1, min_row - 1, max_col - 1, max_row - 1)


def generate_full_styled_pdf(values, styles, merged_ranges, col_widths, pdf_path, title="Combined DATA Sheets"):
    """Generate a PDF preserving colors, fonts, alignment, merges, and column widths."""
    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4))
    elements = []
    stylesheets = getSampleStyleSheet()

    elements.append(Paragraph(title, stylesheets["Title"]))
    elements.append(Spacer(1, 12))

    n_rows = len(values)
    n_cols = len(values[0]) if n_rows else 0

    table = Table(values, repeatRows=1, colWidths=col_widths[:n_cols])
    tstyle = TableStyle([('GRID', (0,0), (-1,-1), 0.25, colors.grey)])

    # --- Apply merged cells ---
    for r in merged_ranges:
        c1, r1, c2, r2 = excel_range_to_indices(r)
        tstyle.add('SPAN', (c1, r1), (c2, r2))

    # --- Apply per-cell styles ---
    for r in range(n_rows):
        for c in range(n_cols):
            cell_style = styles[r][c]
            # Background
            if cell_style["bg_color"]:
                try:
                    tstyle.add('BACKGROUND', (c, r), (c, r), colors.HexColor(cell_style["bg_color"]))
                except:
                    pass
            # Bold / Italic
            if cell_style["bold"]:
                tstyle.add('FONTNAME', (c, r), (c, r), 'Helvetica-Bold')
            elif cell_style["italic"]:
                tstyle.add('FONTNAME', (c, r), (c, r), 'Helvetica-Oblique')
            # Alignment
            if cell_style["align"] in ['LEFT', 'CENTER', 'RIGHT']:
                tstyle.add('ALIGN', (c, r), (c, r), cell_style["align"])

    # --- Header styling ---
    tstyle.add('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey)
    tstyle.add('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
    tstyle.add('ALIGN', (0, 0), (-1, 0), 'CENTER')

    table.setStyle(tstyle)
    elements.append(table)
    doc.build(elements)


def process_xlsm_files_fully_styled(xlsm_files, output_pdf="combined_fully_styled.pdf"):
    """Main entry point: read XLSM DATA sheets and create styled PDF."""
    values, styles, merges, widths = combine_multiple_xlsm(xlsm_files)
    generate_full_styled_pdf(values, styles, merges, widths, output_pdf)
    print(f"✅ Fully styled PDF generated: {output_pdf}")


# Example:
# files = ["report1.xlsm", "report2.xlsm"]
# process_xlsm_files_fully_styled(files, "output_final.pdf")
